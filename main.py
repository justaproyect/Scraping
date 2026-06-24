"""
ENERGY SCRAPER - Orquestador Principal
Busca empresas de sistemas energéticos / ahorro de energía
en Google Places + directorios web, las enriquece y guarda en Firebase
"""

import os
import sys
import json
import time
from datetime import datetime
from typing import List, Dict

import config
from google_places_scraper import GooglePlacesScraper
from web_scraper import WebScraper
from firebase_db import FirebaseDB


class EnergyCompanyScraper:
    def __init__(self):
        self.places_scraper = GooglePlacesScraper()
        self.web_scraper = WebScraper()
        self.firebase = FirebaseDB()
        self.all_companies = []

    def run(self):
        """Ejecuta el proceso completo de recolección."""
        print("=" * 70)
        print("ENERGY COMPANY SCRAPER - Recolección de Empresas")
        print(f"Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 70)

        # Paso 1: Buscar en Google Places
        print("\n[ PASO 1 ] Buscando en Google Places API...")
        places_companies = self.places_scraper.run_all_searches()
        self.all_companies.extend(places_companies)

        # Paso 2: Buscar en directorios web
        print("\n[ PASO 2 ] Buscando en directorios web...")
        web_companies = self._search_directories()
        self.all_companies.extend(web_companies)

        # Paso 3: Eliminar duplicados internos
        print("\n[ PASO 3 ] Eliminando duplicados...")
        self.all_companies = self._deduplicate(self.all_companies)

        # Paso 4: Enriquecer con datos de sitios web
        print("\n[ PASO 4 ] Enriqueciendo datos desde sitios web...")
        self._enrich_from_websites()

        # Paso 5: Guardar en Firebase
        print("\n[ PASO 5 ] Guardando en Firebase Firestore...")
        self._save_to_firebase()

        # Paso 6: Exportar respaldo a Excel
        if config.EXPORT_TO_EXCEL:
            print("\n[ PASO 6 ] Exportando respaldo a Excel...")
            self._export_to_excel()

        # Resumen final
        print("\n" + "=" * 70)
        print("RESUMEN FINAL")
        print(f"  Total empresas recolectadas: {len(self.all_companies)}")
        if self.firebase._connected:
            print(f"  Total en Firebase Firestore: {self.firebase.count_companies()}")
        print(f"  Fin: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 70)

    def _search_directories(self) -> List[Dict]:
        """Busca empresas en directorios web."""
        results = []
        for directory in config.DIRECTORIES:
            for term in config.SEARCH_TERMS[:5]:  # Solo primeros términos para no abusar
                print(f"\n  Directorio: {directory} | Término: '{term}'")
                companies = self.web_scraper.search_directory(
                    f"https://{directory}", term,
                    max_pages=config.MAX_DIRECTORY_PAGES
                )
                for c in companies:
                    c["termino_busqueda"] = term
                results.extend(companies)
                time.sleep(config.REQUEST_DELAY)

        print(f"\n  -> {len(results)} empresas encontradas en directorios")
        return results

    def _deduplicate(self, companies: List[Dict]) -> List[Dict]:
        """Elimina empresas duplicadas por nombre + teléfono."""
        seen = set()
        unique = []
        for c in companies:
            nombre = c.get("nombre", "").strip().lower()[:30]
            telefono = c.get("telefono", "").strip()[:10]
            key = f"{nombre}|{telefono}"
            if key not in seen:
                seen.add(key)
                unique.append(c)
        return unique

    def _enrich_from_websites(self):
        """Enriquece empresas con datos de su sitio web."""
        enriched = 0
        for i, company in enumerate(self.all_companies):
            website = company.get("sitio_web", "")
            if website and not company.get("telefono"):
                print(f"  [{i+1}/{len(self.all_companies)}] {company['nombre'][:40]}...")
                company = self.web_scraper.enrich_from_website(company)
                self.all_companies[i] = company
                enriched += 1
                time.sleep(config.REQUEST_DELAY)

                # Pausa cada 10 para no saturar
                if enriched > 0 and enriched % 10 == 0:
                    print(f"  Pausa breve... ({enriched} enriquecidas)")

        print(f"  -> {enriched} empresas enriquecidas con datos web")

    def _save_to_firebase(self):
        """Guarda todas las empresas en Firebase."""
        if not self.firebase.connect():
            print("  [!] No se pudo conectar a Firebase. Los datos solo están en memoria.")
            return

        inserted, updated, errors = self.firebase.save_company_batch(self.all_companies)
        print(f"  -> Insertadas: {inserted}")
        print(f"  -> Actualizadas: {updated}")
        print(f"  -> Errores: {errors}")

    def _export_to_excel(self):
        """Exporta los datos a un archivo Excel como respaldo."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Empresas Energéticas"

            # Encabezados
            headers = [
                "Nombre", "Teléfono", "Teléfonos Adicionales", "Email",
                "Emails Adicionales", "Sitio Web", "Dirección",
                "Calificación", "Reseñas", "Categorías",
                "Término Búsqueda", "Ubicación", "Fuente"
            ]

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Datos
            for row, company in enumerate(self.all_companies, 2):
                ws.cell(row=row, column=1, value=company.get("nombre", ""))
                ws.cell(row=row, column=2, value=company.get("telefono", ""))
                phones = company.get("telefonos_adicionales", [])
                ws.cell(row=row, column=3, value="; ".join(phones[:5]) if phones else "")
                ws.cell(row=row, column=4, value=company.get("email", ""))
                emails = company.get("emails_adicionales", [])
                ws.cell(row=row, column=5, value="; ".join(emails[:5]) if emails else "")
                ws.cell(row=row, column=6, value=company.get("sitio_web", ""))
                ws.cell(row=row, column=7, value=company.get("direccion", ""))
                ws.cell(row=row, column=8, value=company.get("calificacion", ""))
                ws.cell(row=row, column=9, value=company.get("total_resenas", ""))
                ws.cell(row=row, column=10, value=company.get("categorias", ""))
                ws.cell(row=row, column=11, value=company.get("termino_busqueda", ""))
                ws.cell(row=row, column=12, value=company.get("ubicacion", ""))
                ws.cell(row=row, column=13, value=company.get("fuente", ""))

            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 25

            output_path = config.EXCEL_OUTPUT_PATH
            wb.save(output_path)
            print(f"  -> Excel guardado: {output_path}")

        except ImportError:
            print("  [!] openpyxl no instalado. No se generó Excel.")
        except Exception as e:
            print(f"  [!] Error exportando Excel: {e}")


def main():
    scraper = EnergyCompanyScraper()
    scraper.run()


if __name__ == "__main__":
    main()
