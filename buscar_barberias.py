import requests
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

API_KEY = "AIzaSyCxfBhccJevL5NDSL-mPYSTqvjA7gHVGDk"
LOCATION = "Barranquilla"

TERMS = ["barberias", "barber shop", "barberia"]

def text_search(term):
    params = {"query": f"{term} en {LOCATION}", "key": API_KEY, "language": "es", "type": "establishment"}
    try:
        resp = requests.get("https://maps.googleapis.com/maps/api/place/textsearch/json", params=params, timeout=15)
        data = resp.json()
        if data.get("status") != "OK":
            return []
        return data.get("results", [])
    except:
        return []

def get_detail(place_id):
    params = {
        "place_id": place_id, "key": API_KEY, "language": "es",
        "fields": "name,formatted_phone_number,international_phone_number,formatted_address,website,rating,user_ratings_total,types",
    }
    try:
        resp = requests.get("https://maps.googleapis.com/maps/api/place/details/json", params=params, timeout=10)
        data = resp.json()
        if data.get("status") != "OK":
            return None
        r = data.get("result", {})
        phone = r.get("formatted_phone_number") or r.get("international_phone_number") or ""
        return {
            "nombre": r.get("name", ""),
            "telefono": phone,
            "direccion": r.get("formatted_address", ""),
            "sitio_web": r.get("website", ""),
            "calificacion": r.get("rating") if r.get("rating") is not None else "",
            "votos": r.get("user_ratings_total", ""),
        }
    except:
        return None

print(f"[{datetime.now().strftime('%H:%M:%S')}] Buscando barberias en {LOCATION}...")

# Fase 1
all_places = {}
for i, term in enumerate(TERMS, 1):
    print(f"  [{i}/{len(TERMS)}] '{term}'...")
    results = text_search(term)
    for place in results:
        pid = place.get("place_id")
        if pid and pid not in all_places:
            all_places[pid] = place.get("name", "")

print(f"\n  Barbers unicas encontradas: {len(all_places)}")

# Fase 2
print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Obteniendo telefonos...")
barbers = []
done = 0
for pid, name in all_places.items():
    done += 1
    if done % 10 == 0:
        print(f"  {done}/{len(all_places)}")
    detail = get_detail(pid)
    if detail:
        barbers.append(detail)
    else:
        barbers.append({"nombre": name, "telefono": "", "direccion": "", "sitio_web": "", "calificacion": "", "votos": 0})
    time.sleep(0.3)

with_phone = [b for b in barbers if b.get("telefono")]
without_phone = [b for b in barbers if not b.get("telefono")]

print(f"\n{'='*60}")
print(f"BARBERIAS EN {LOCATION}")
print(f"{'='*60}")
print(f"Total:       {len(barbers)}")
print(f"Con telefono: {len(with_phone)}")
print(f"Sin telefono: {len(without_phone)}")

if with_phone:
    print(f"\n--- BARBERIAS CON TELEFONO ({len(with_phone)}) ---")
    for b in sorted(with_phone, key=lambda x: x.get("nombre")):
        nombre = b.get("nombre", "")
        telefono = b.get("telefono", "")
        rating = b.get("calificacion", "")
        print(f"  {nombre[:45]:45s} | {telefono:20s} | {'' if rating is None else rating}")

# Guardar
output = "barberias_barranquilla.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Barberias Barranquilla"

headers = ["Nombre", "Telefono", "Direccion", "Sitio Web", "Calificacion", "Votos"]
hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = hf; c.font = hfont

for row, b in enumerate(barbers, 2):
    ws.cell(row=row, column=1, value=b.get("nombre", ""))
    ws.cell(row=row, column=2, value=b.get("telefono", ""))
    ws.cell(row=row, column=3, value=b.get("direccion", ""))
    ws.cell(row=row, column=4, value=b.get("sitio_web", ""))
    ws.cell(row=row, column=5, value=b.get("calificacion", ""))
    ws.cell(row=row, column=6, value=b.get("votos", ""))

for col in range(1, 7):
    ws.column_dimensions[chr(64 + col)].width = 35

wb.save(output)
print(f"\nExcel guardado: {output}")
