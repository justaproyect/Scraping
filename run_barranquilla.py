import requests
import time
import json
import re
from datetime import datetime

API_KEY = "AIzaSyCxfBhccJevL5NDSL-mPYSTqvjA7gHVGDk"

TERMS = [
    "sistemas energeticos", "ahorro de energia", "paneles solares",
    "energia solar", "eficiencia energetica", "energias renovables",
    "instalacion paneles solares", "sistemas fotovoltaicos",
    "auditoria energetica", "consultoria energetica"
]

LOCATION = "Barranquilla"

def text_search(term):
    """Busqueda rapida solo text search (sin detail API)."""
    params = {"query": f"{term} en {LOCATION}", "key": API_KEY, "language": "es"}
    resp = requests.get(
        "https://maps.googleapis.com/maps/api/place/textsearch/json",
        params=params, timeout=15
    )
    data = resp.json()
    if data.get("status") != "OK":
        return []
    return data.get("results", [])

def get_detail(place_id):
    """Obtiene detalle con telefono."""
    params = {
        "place_id": place_id, "key": API_KEY, "language": "es",
        "fields": "name,formatted_phone_number,international_phone_number,formatted_address,website,rating,user_ratings_total,types",
    }
    try:
        resp = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params=params, timeout=10
        )
        data = resp.json()
        if data.get("status") != "OK":
            return None
        r = data.get("result", {})
        phone = r.get("formatted_phone_number") or r.get("international_phone_number") or ""
        return {
            "nombre": r.get("name", ""),
            "telefono": phone,
            "direccion": r.get("formatted_address", ""),
            "sitio_web": r.get("website", ""),
            "calificacion": r.get("rating"),
        }
    except:
        return None

# FASE 1: Buscar todos los place_ids rapidamente
print(f"[{datetime.now().strftime('%H:%M:%S')}] Fase 1: Buscando en Google Places...")
all_places = {}  # place_id -> nombre
for i, term in enumerate(TERMS, 1):
    print(f"  [{i}/{len(TERMS)}] {term}...")
    results = text_search(term)
    for place in results:
        pid = place.get("place_id")
        if pid and pid not in all_places:
            all_places[pid] = place.get("name", "")

print(f"\n  Total places unicos encontrados: {len(all_places)}")

# FASE 2: Obtener detalle (con telefono) de cada uno
print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Fase 2: Obteniendo telefonos...")
companies = []
done = 0
for pid, name in all_places.items():
    done += 1
    if done % 5 == 0:
        print(f"  {done}/{len(all_places)} - ultimo: {name[:40]}")
    
    detail = get_detail(pid)
    if detail:
        companies.append(detail)
    else:
        companies.append({"nombre": name, "telefono": "", "direccion": "", "sitio_web": "", "calificacion": ""})
    time.sleep(0.3)

# Resultados
with_phone = [c for c in companies if c.get("telefono")]
without_phone = [c for c in companies if not c.get("telefono")]

print(f"\n{'='*60}")
print(f"RESULTADOS - {LOCATION}")
print(f"{'='*60}")
print(f"Total empresas: {len(companies)}")
print(f"Con telefono:   {len(with_phone)}")
print(f"Sin telefono:   {len(without_phone)}")

if with_phone:
    print(f"\n--- EMPRESAS CON TELEFONO ({len(with_phone)}) ---")
    for c in sorted(with_phone, key=lambda x: x.get("nombre")):
        nombre = c.get("nombre", "")
        telefono = c.get("telefono", "")
        web = c.get("sitio_web", "")
        print(f"  {nombre[:45]:45s} | {telefono:20s} | {web[:30]}")

# Guardar Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    output = f"empresas_{LOCATION.lower()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = LOCATION
    
    headers = ["Nombre", "Telefono", "Direccion", "Sitio Web", "Calificacion"]
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont
    
    for row, c in enumerate(companies, 2):
        ws.cell(row=row, column=1, value=c.get("nombre", ""))
        ws.cell(row=row, column=2, value=c.get("telefono", ""))
        ws.cell(row=row, column=3, value=c.get("direccion", ""))
        ws.cell(row=row, column=4, value=c.get("sitio_web", ""))
        ws.cell(row=row, column=5, value=c.get("calificacion", ""))
    
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 35
    
    wb.save(output)
    print(f"\nExcel guardado: {output}")
except Exception as e:
    print(f"\nError guardando Excel: {e}")
