import requests
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

API_KEY = "AIzaSyCxfBhccJevL5NDSL-mPYSTqvjA7gHVGDk"
LOCATION = "Barranquilla"

# Tipos de clientes objetivo (negocios que NECESITAN los servicios)
BUSINESS_TYPES = [
    "hoteles",
    "moteles",
    "centros comerciales",
    "restaurantes",
    "clinicas",
    "colegios",
    "plazoletas de comidas",
    "grandes empresas",
    "almacenes grandes",
    "gimnasios",
    "hospitales",
    "universidades",
    "bancos",
    "supermercados",
    "concesionarios",
    "bodegas industriales",
    "centros empresariales",
    "edificios de oficinas",
    "industrias",
    "fabricas",
]

def text_search(term):
    params = {"query": f"{term} en {LOCATION}", "key": API_KEY, "language": "es", "type": "establishment"}
    try:
        resp = requests.get("https://maps.googleapis.com/maps/api/place/textsearch/json", params=params, timeout=15)
        data = resp.json()
        if data.get("status") != "OK":
            return []
        return data.get("results", [])
    except:
        return []

def get_detail(place_id):
    params = {
        "place_id": place_id, "key": API_KEY, "language": "es",
        "fields": "name,formatted_phone_number,international_phone_number,formatted_address,website,rating,user_ratings_total,types",
    }
    try:
        resp = requests.get("https://maps.googleapis.com/maps/api/place/details/json", params=params, timeout=10)
        data = resp.json()
        if data.get("status") != "OK":
            return None
        r = data.get("result", {})
        phone = r.get("formatted_phone_number") or r.get("international_phone_number") or ""
        return {
            "nombre": r.get("name", ""),
            "telefono": phone,
            "direccion": r.get("formatted_address", ""),
            "sitio_web": r.get("website", ""),
            "calificacion": r.get("rating"),
            "tipo_busqueda": "",
        }
    except:
        return None

# FASE 1: Buscar todos los lugares
print(f"[{datetime.now().strftime('%H:%M:%S')}] Buscando tipos de negocio en {LOCATION}...")
all_places = {}  # place_id -> (nombre, tipo_busqueda)

for i, btype in enumerate(BUSINESS_TYPES, 1):
    print(f"  [{i}/{len(BUSINESS_TYPES)}] {btype}...")
    results = text_search(btype)
    for place in results:
        pid = place.get("place_id")
        if pid and pid not in all_places:
            all_places[pid] = (place.get("name", ""), btype)

print(f"\n  Total lugares unicos encontrados: {len(all_places)}")

# FASE 2: Obtener detalle con telefono
print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Obteniendo telefonos...")
companies = []
done = 0
for pid, (name, btype) in all_places.items():
    done += 1
    if done % 10 == 0:
        print(f"  {done}/{len(all_places)}")
    
    detail = get_detail(pid)
    if detail:
        detail["tipo_busqueda"] = btype
        companies.append(detail)
    else:
        companies.append({"nombre": name, "telefono": "", "direccion": "", "sitio_web": "", "calificacion": "", "tipo_busqueda": btype})
    time.sleep(0.3)

# RESULTADOS
with_phone = [c for c in companies if c.get("telefono")]
without_phone = [c for c in companies if not c.get("telefono")]

print(f"\n{'='*70}")
print(f"RESULTADOS - CLIENTES POTENCIALES EN {LOCATION}")
print(f"{'='*70}")
print(f"Total negocios encontrados: {len(companies)}")
print(f"Con telefono:              {len(with_phone)}")
print(f"Sin telefono:              {len(without_phone)}")

# Por tipo de negocio
from collections import Counter
type_counts = Counter(c.get("tipo_busqueda", "sin_tipo") for c in companies)
print(f"\n--- POR TIPO DE NEGOCIO ---")
for t, count in sorted(type_counts.items()):
    print(f"  {t:30s} {count}")

if with_phone:
    print(f"\n--- LISTA COMPLETA CON TELEFONO ---")
    for c in sorted(with_phone, key=lambda x: x.get("tipo_busqueda", "")):
        nombre = c.get("nombre", "")
        telefono = c.get("telefono", "")
        tipo = c.get("tipo_busqueda", "")
        print(f"  [{tipo:25s}] {nombre[:40]:40s} | {telefono:20s}")

# Guardar Excel
output = f"clientes_{LOCATION.lower()}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Clientes Potenciales"

headers = ["Nombre", "Telefono", "Direccion", "Sitio Web", "Calificacion", "Tipo Negocio", "Servicios Recomendados"]
hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = hf; c.font = hfont

# Mapa de servicios recomendados por tipo de negocio
recomendaciones = {
    "hoteles": "Climatización, Instalaciones Eléctricas, CCTV, Energía Solar, SST",
    "moteles": "Instalaciones Eléctricas, Climatización, CCTV, Control de Acceso",
    "centros comerciales": "Instalaciones Eléctricas, Climatización, CCTV, Energía Solar, SST",
    "restaurantes": "Instalaciones Eléctricas, Climatización, CCTV",
    "clinicas": "Instalaciones Eléctricas, Climatización, CCTV, SST, UPS",
    "colegios": "Instalaciones Eléctricas, Climatización, CCTV, SST, Energía Solar",
    "plazoletas de comidas": "Instalaciones Eléctricas, Climatización, CCTV",
    "grandes empresas": "Instalaciones Eléctricas, Climatización, CCTV, Energía Solar, SST, Control de Acceso",
    "almacenes grandes": "CCTV, Instalaciones Eléctricas, Climatización, Control de Acceso",
    "gimnasios": "Climatización, Instalaciones Eléctricas, CCTV",
    "hospitales": "Instalaciones Eléctricas, Climatización, SST, CCTV, UPS, Energía Solar",
    "universidades": "Instalaciones Eléctricas, Climatización, CCTV, SST, Energía Solar, Control de Acceso",
    "bancos": "CCTV, Instalaciones Eléctricas, UPS, Control de Acceso",
    "supermercados": "Instalaciones Eléctricas, Climatización, CCTV, Energía Solar",
    "concesionarios": "Instalaciones Eléctricas, Climatización, CCTV, Energía Solar",
    "bodegas industriales": "Instalaciones Eléctricas, CCTV, SST, Energía Solar",
    "centros empresariales": "Instalaciones Eléctricas, Climatización, CCTV, Control de Acceso, SST",
    "edificios de oficinas": "Instalaciones Eléctricas, Climatización, CCTV, Control de Acceso",
    "industrias": "Instalaciones Eléctricas, SST, CCTV, Energía Solar",
    "fabricas": "Instalaciones Eléctricas, SST, CCTV, Energía Solar",
}

for row, c in enumerate(companies, 2):
    tipo = c.get("tipo_busqueda", "")
    ws.cell(row=row, column=1, value=c.get("nombre", ""))
    ws.cell(row=row, column=2, value=c.get("telefono", ""))
    ws.cell(row=row, column=3, value=c.get("direccion", ""))
    ws.cell(row=row, column=4, value=c.get("sitio_web", ""))
    ws.cell(row=row, column=5, value=c.get("calificacion", ""))
    ws.cell(row=row, column=6, value=tipo)
    ws.cell(row=row, column=7, value=recomendaciones.get(tipo, "Instalaciones Eléctricas, Climatización, CCTV"))

for col in range(1, 8):
    ws.column_dimensions[chr(64 + col)].width = 35

wb.save(output)
print(f"\nExcel guardado: {output}")
print(f"[{datetime.now().strftime('%H:%M:%S')}] FIN")
