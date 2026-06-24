import json, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE, "datos")
CAT_DIR = os.path.join(DATA_DIR, "por_categoria")
os.makedirs(CAT_DIR, exist_ok=True)

with open(os.path.join(DATA_DIR, "enviar_barranquilla.json"), "r", encoding="utf-8") as f:
    negocios = json.load(f)

def limpiar_numero(raw):
    num = re.sub(r'\s*(ext|x|opción)\s*[.\s]*\d+\s*$', '', raw, flags=re.I)
    num = re.sub(r'\D', '', num)
    if num.startswith("57"): num = num[2:]
    if num.startswith("0"): num = num[1:]
    if not num.startswith("3") or len(num) != 10: return None
    return "57" + num

por_tipo = defaultdict(list)
for n in negocios:
    por_tipo[n.get("tipo", "sin_tipo")].append(n)

header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
cell_font = Font(color="CCCCCC", size=10)
dim_font = Font(color="666666", size=10)
sep_font = Font(color="666666", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin', color='333333'),
    right=Side(style='thin', color='333333'),
    top=Side(style='thin', color='333333'),
    bottom=Side(style='thin', color='333333'),
)

total_con = 0
total_sin = 0

for tipo in sorted(por_tipo.keys()):
    items = por_tipo[tipo]
    con = [n for n in items if n.get("telefono") and limpiar_numero(n.get("telefono", ""))]
    sin = [n for n in items if not n.get("telefono") or not limpiar_numero(n.get("telefono", ""))]
    total_con += len(con)
    total_sin += len(sin)

    wb = Workbook()
    ws = wb.active
    ws.title = tipo[:31]

    headers = ["Nombre", "Telefono", "WhatsApp", "Direccion", "Servicios", "Mensaje"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = thin_border

    row = 2
    for n in con:
        tel = n.get("telefono", "")
        num = limpiar_numero(tel)
        wa = f"https://wa.me/{num}" if num else ""
        svc = ", ".join(n.get("servicios_recomendados", [])[:3])
        msg = n.get("mensaje", "")
        vals = [n.get("nombre", ""), tel, wa, n.get("direccion", ""), svc, msg]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = cell_font; c.border = thin_border
            c.alignment = Alignment(vertical='top', wrap_text=(col == 6))
        row += 1

    if sin:
        row += 1
        c = ws.cell(row=row, column=1, value=f"SIN WHATSAPP ({len(sin)})")
        c.font = sep_font
        row += 1
        for n in sin:
            vals = [n.get("nombre", ""), n.get("telefono", ""), "", n.get("direccion", "")]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = dim_font; c.border = thin_border
            row += 1

    widths = [40, 18, 50, 40, 45, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    filename = f"{tipo.replace(' ', '_').lower()}.xlsx"
    wb.save(os.path.join(CAT_DIR, filename))
    print(f"  {tipo:30s} -> {filename} ({len(con)} WA, {len(sin)} sin)")

print(f"\nTotal: {total_con} con WhatsApp, {total_sin} sin WhatsApp")
print(f"Carpeta: {CAT_DIR}")
